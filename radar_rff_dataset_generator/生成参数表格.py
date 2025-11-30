"""
生成雷达RFF数据集参数Excel表格

包含：
1. 20类设备的详细RFF参数
2. 26种调制方式列表
3. 520个组合类别映射表
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from config_rff_parameters import generate_device_parameters, MODULATION_LIST, get_combined_class_id
import numpy as np

def create_excel_with_parameters(output_file: str = "雷达RFF数据集参数表.xlsx"):
    """创建包含所有参数的Excel文件"""
    
    # 生成设备参数（使用diverse=True，便于区分）
    devices = generate_device_parameters(num_devices=20, seed=42, diverse=True)
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    
    # 删除默认的Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ============================================================================
    # 工作表1: 20类设备参数详情
    # ============================================================================
    ws_devices = wb.create_sheet("设备参数详情", 0)
    
    # 设置列宽
    ws_devices.column_dimensions['A'].width = 5
    ws_devices.column_dimensions['B'].width = 20
    ws_devices.column_dimensions['C'].width = 18
    ws_devices.column_dimensions['D'].width = 18
    ws_devices.column_dimensions['E'].width = 18
    ws_devices.column_dimensions['F'].width = 18
    ws_devices.column_dimensions['G'].width = 15
    ws_devices.column_dimensions['H'].width = 15
    ws_devices.column_dimensions['I'].width = 15
    ws_devices.column_dimensions['J'].width = 12
    ws_devices.column_dimensions['K'].width = 12
    
    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题行
    headers = [
        "ID", "设备名称", 
        "IQ幅度不平衡\n(0.95-1.05)", 
        "IQ相位不平衡\n(-5° ~ +5°)", 
        "载波频偏\n(-50 ~ +50 ppm)", 
        "采样率偏移\n(-20 ~ +20 ppm)", 
        "相位噪声\n(0.001-0.01)", 
        "功放平滑因子\n(1.0-5.0)", 
        "功放饱和电平\n(0.7-1.2)", 
        "DC偏移_I\n(-0.05~+0.05)", 
        "DC偏移_Q\n(-0.05~+0.05)"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws_devices.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 设置标题行高度
    ws_devices.row_dimensions[1].height = 40
    
    # 填充设备数据
    for idx, device in enumerate(devices, start=2):
        ws_devices.cell(row=idx, column=1, value=device.device_id).alignment = cell_alignment
        ws_devices.cell(row=idx, column=2, value=device.device_name).alignment = cell_alignment
        ws_devices.cell(row=idx, column=3, value=round(device.iq_imbalance_amplitude, 4)).alignment = cell_alignment
        ws_devices.cell(row=idx, column=4, value=round(device.iq_imbalance_phase_deg, 2)).alignment = cell_alignment
        ws_devices.cell(row=idx, column=5, value=round(device.carrier_freq_offset_ppm, 2)).alignment = cell_alignment
        ws_devices.cell(row=idx, column=6, value=round(device.sampling_rate_offset_ppm, 2)).alignment = cell_alignment
        ws_devices.cell(row=idx, column=7, value=round(device.phase_noise_level, 4)).alignment = cell_alignment
        ws_devices.cell(row=idx, column=8, value=round(device.pa_rapp_smoothness, 2)).alignment = cell_alignment
        ws_devices.cell(row=idx, column=9, value=round(device.pa_saturation_level, 2)).alignment = cell_alignment
        ws_devices.cell(row=idx, column=10, value=round(device.dc_offset_i, 4)).alignment = cell_alignment
        ws_devices.cell(row=idx, column=11, value=round(device.dc_offset_q, 4)).alignment = cell_alignment
        
        # 添加边框
        for col in range(1, 12):
            ws_devices.cell(row=idx, column=col).border = border
    
    # ============================================================================
    # 工作表2: 26种调制方式
    # ============================================================================
    ws_modulations = wb.create_sheet("调制方式列表", 1)
    
    # 设置列宽
    ws_modulations.column_dimensions['A'].width = 10
    ws_modulations.column_dimensions['B'].width = 15
    ws_modulations.column_dimensions['C'].width = 25
    ws_modulations.column_dimensions['D'].width = 40
    
    # 标题行
    mod_headers = ["ID", "调制类型", "调制系列", "说明"]
    for col, header in enumerate(mod_headers, start=1):
        cell = ws_modulations.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    ws_modulations.row_dimensions[1].height = 25
    
    # 调制系列分类
    modulation_info = {
        "16qam": ("QAM系列", "16阶正交幅度调制"),
        "64qam": ("QAM系列", "64阶正交幅度调制"),
        "256qam": ("QAM系列", "256阶正交幅度调制"),
        "1024qam": ("QAM系列", "1024阶正交幅度调制"),
        
        "bpsk": ("PSK系列", "二进制相移键控"),
        "qpsk": ("PSK系列", "四相相移键控"),
        "8psk": ("PSK系列", "8相相移键控"),
        "16psk": ("PSK系列", "16相相移键控"),
        "32psk": ("PSK系列", "32相相移键控"),
        
        "2fsk": ("FSK系列", "二进制频移键控"),
        "4fsk": ("FSK系列", "4进制频移键控"),
        "8fsk": ("FSK系列", "8进制频移键控"),
        "16fsk": ("FSK系列", "16进制频移键控"),
        
        "2gfsk": ("GFSK系列", "二进制高斯频移键控"),
        "4gfsk": ("GFSK系列", "4进制高斯频移键控"),
        "8gfsk": ("GFSK系列", "8进制高斯频移键控"),
        "16gfsk": ("GFSK系列", "16进制高斯频移键控"),
        
        "2msk": ("MSK系列", "二进制最小频移键控"),
        "4msk": ("MSK系列", "4进制最小频移键控"),
        "8msk": ("MSK系列", "8进制最小频移键控"),
        "16msk": ("MSK系列", "16进制最小频移键控"),
        
        "am-dsb": ("AM系列", "双边带调幅"),
        "am-dsb-sc": ("AM系列", "双边带抑制载波调幅"),
        "am-lsb": ("AM系列", "下边带调幅"),
        "am-usb": ("AM系列", "上边带调幅"),
        
        "fm": ("FM系列", "频率调制"),
    }
    
    # 填充调制数据
    series_colors = {
        "QAM系列": "E7E6FF",
        "PSK系列": "FFE6E6",
        "FSK系列": "E6F7FF",
        "GFSK系列": "E6FFE6",
        "MSK系列": "FFF4E6",
        "AM系列": "FFE6F0",
        "FM系列": "F0F0F0",
    }
    
    for idx, mod in enumerate(MODULATION_LIST, start=2):
        series, description = modulation_info[mod]
        
        ws_modulations.cell(row=idx, column=1, value=idx-2).alignment = cell_alignment
        ws_modulations.cell(row=idx, column=2, value=mod.upper()).alignment = cell_alignment
        ws_modulations.cell(row=idx, column=3, value=series).alignment = cell_alignment
        ws_modulations.cell(row=idx, column=4, value=description).alignment = Alignment(horizontal="left", vertical="center")
        
        # 根据系列添加背景色
        fill_color = PatternFill(start_color=series_colors[series], 
                                 end_color=series_colors[series], 
                                 fill_type="solid")
        for col in range(1, 5):
            cell = ws_modulations.cell(row=idx, column=col)
            cell.fill = fill_color
            cell.border = border
    
    # ============================================================================
    # 工作表3: 设备-调制组合映射表（520类）
    # ============================================================================
    ws_combinations = wb.create_sheet("设备调制组合映射", 2)
    
    # 设置列宽
    ws_combinations.column_dimensions['A'].width = 12
    ws_combinations.column_dimensions['B'].width = 12
    ws_combinations.column_dimensions['C'].width = 20
    ws_combinations.column_dimensions['D'].width = 15
    ws_combinations.column_dimensions['E'].width = 15
    
    # 标题行
    comb_headers = ["组合ID", "设备ID", "设备名称", "调制ID", "调制类型"]
    for col, header in enumerate(comb_headers, start=1):
        cell = ws_combinations.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    ws_combinations.row_dimensions[1].height = 25
    
    # 填充组合数据
    for device in devices:
        for mod_id, mod_name in enumerate(MODULATION_LIST):
            combined_id = get_combined_class_id(device.device_id, mod_id)
            row = combined_id + 2  # +2 because header is row 1
            
            ws_combinations.cell(row=row, column=1, value=combined_id).alignment = cell_alignment
            ws_combinations.cell(row=row, column=2, value=device.device_id).alignment = cell_alignment
            ws_combinations.cell(row=row, column=3, value=device.device_name).alignment = cell_alignment
            ws_combinations.cell(row=row, column=4, value=mod_id).alignment = cell_alignment
            ws_combinations.cell(row=row, column=5, value=mod_name.upper()).alignment = cell_alignment
            
            # 添加边框
            for col in range(1, 6):
                ws_combinations.cell(row=row, column=col).border = border
            
            # 每个设备使用不同的背景色（淡色）
            device_colors = [
                "FFE6E6", "E6F7FF", "E6FFE6", "FFF4E6", "FFE6F0",
                "F0E6FF", "E6F0FF", "F0FFE6", "FFE6CC", "E6E6FF",
                "FFE6E6", "E6F7FF", "E6FFE6", "FFF4E6", "FFE6F0",
                "F0E6FF", "E6F0FF", "F0FFE6", "FFE6CC", "E6E6FF",
            ]
            fill_color = PatternFill(start_color=device_colors[device.device_id], 
                                     end_color=device_colors[device.device_id], 
                                     fill_type="solid")
            for col in range(1, 6):
                ws_combinations.cell(row=row, column=col).fill = fill_color
    
    # ============================================================================
    # 工作表4: 参数说明
    # ============================================================================
    ws_info = wb.create_sheet("参数说明", 3)
    
    # 设置列宽
    ws_info.column_dimensions['A'].width = 25
    ws_info.column_dimensions['B'].width = 80
    
    # 标题
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    
    info_data = [
        ("", "雷达射频指纹（RFF）数据集参数说明", title_fill, title_font),
        ("", "", None, None),
        ("数据集规模", "", header_fill, header_font),
        ("设备数量", "20个不同的雷达设备", None, None),
        ("调制类型", "26种调制方式", None, None),
        ("组合类别", "520个（设备-调制）组合", None, None),
        ("每类样本数", "3,000个", None, None),
        ("总样本数", "1,560,000个", None, None),
        ("数据集大小", "约 48.6 GB", None, None),
        ("", "", None, None),
        ("RFF参数说明", "", header_fill, header_font),
        ("IQ幅度不平衡", "I/Q两路信号的幅度差异，范围：0.95-1.05（±5%）", None, None),
        ("IQ相位不平衡", "I/Q两路信号的相位差异，范围：-5° ~ +5°", None, None),
        ("载波频偏 (CFO)", "载波频率偏移，单位：ppm（百万分之一），范围：-50 ~ +50 ppm", None, None),
        ("采样率偏移 (SRO)", "采样时钟偏移，单位：ppm，范围：-20 ~ +20 ppm", None, None),
        ("相位噪声", "振荡器相位抖动强度，范围：0.001-0.01", None, None),
        ("功放平滑因子", "Rapp模型平滑参数，控制非线性失真程度，范围：1.0-5.0", None, None),
        ("功放饱和电平", "功率放大器饱和点，范围：0.7-1.2", None, None),
        ("DC偏移_I", "I通道直流偏移，范围：-0.05 ~ +0.05", None, None),
        ("DC偏移_Q", "Q通道直流偏移，范围：-0.05 ~ +0.05", None, None),
        ("", "", None, None),
        ("调制系列", "", header_fill, header_font),
        ("QAM系列", "4种：16QAM, 64QAM, 256QAM, 1024QAM", None, None),
        ("PSK系列", "5种：BPSK, QPSK, 8PSK, 16PSK, 32PSK", None, None),
        ("FSK系列", "4种：2FSK, 4FSK, 8FSK, 16FSK", None, None),
        ("GFSK系列", "4种：2GFSK, 4GFSK, 8GFSK, 16GFSK", None, None),
        ("MSK系列", "4种：2MSK, 4MSK, 8MSK, 16MSK", None, None),
        ("AM系列", "4种：AM-DSB, AM-DSB-SC, AM-LSB, AM-USB", None, None),
        ("FM系列", "1种：FM", None, None),
        ("", "", None, None),
        ("使用说明", "", header_fill, header_font),
        ("标签格式", "每个样本包含两个标签：device_id (0-19) 和 modulation_id (0-25)", None, None),
        ("组合ID计算", "combined_id = device_id × 26 + modulation_id", None, None),
        ("数据格式", "Complex64，形状：(2, 4096)，即 [I, Q] × 4096个采样点", None, None),
        ("存储格式", "Sigmoid WidebandSig53 格式（.sigmf-meta + .sigmf-data）", None, None),
        ("", "", None, None),
        ("⚠️ 重要提示", "", PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"), Font(bold=True, color="FFFFFF", size=12)),
        ("采样率配置", "当前默认配置为 1 MHz，建议修改为 20 MHz 以符合真实场景！", PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid"), Font(bold=True, color="FF0000")),
        ("", "详情请阅读：采样率问题总结.md", None, None),
        ("", "", None, None),
        ("生成日期", "2025-10-18", None, None),
        ("工具", "基于 TorchSig 的雷达RFF数据集生成器", None, None),
    ]
    
    for row_idx, (param, desc, fill, font) in enumerate(info_data, start=1):
        cell_a = ws_info.cell(row=row_idx, column=1, value=param)
        cell_b = ws_info.cell(row=row_idx, column=2, value=desc)
        
        if fill:
            cell_a.fill = fill
            cell_b.fill = fill
        if font:
            cell_a.font = font
            cell_b.font = font
        else:
            cell_a.font = Font(bold=True) if param else Font()
        
        cell_a.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell_b.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # 合并标题行
        if param == "" and desc and (fill == title_fill or "雷达射频指纹" in desc):
            ws_info.merge_cells(f'A{row_idx}:B{row_idx}')
            cell_a.alignment = Alignment(horizontal="center", vertical="center")
    
    # ============================================================================
    # 工作表5: 统计信息
    # ============================================================================
    ws_stats = wb.create_sheet("参数统计", 4)
    
    # 设置列宽
    ws_stats.column_dimensions['A'].width = 25
    ws_stats.column_dimensions['B'].width = 15
    ws_stats.column_dimensions['C'].width = 15
    ws_stats.column_dimensions['D'].width = 15
    ws_stats.column_dimensions['E'].width = 15
    
    # 标题
    stats_headers = ["RFF参数", "最小值", "最大值", "平均值", "标准差"]
    for col, header in enumerate(stats_headers, start=1):
        cell = ws_stats.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 计算统计信息
    params_to_analyze = [
        ("IQ幅度不平衡", [d.iq_imbalance_amplitude for d in devices]),
        ("IQ相位不平衡 (°)", [d.iq_imbalance_phase_deg for d in devices]),
        ("载波频偏 (ppm)", [d.carrier_freq_offset_ppm for d in devices]),
        ("采样率偏移 (ppm)", [d.sampling_rate_offset_ppm for d in devices]),
        ("相位噪声", [d.phase_noise_level for d in devices]),
        ("功放平滑因子", [d.pa_rapp_smoothness for d in devices]),
        ("功放饱和电平", [d.pa_saturation_level for d in devices]),
        ("DC偏移_I", [d.dc_offset_i for d in devices]),
        ("DC偏移_Q", [d.dc_offset_q for d in devices]),
    ]
    
    for row_idx, (param_name, values) in enumerate(params_to_analyze, start=2):
        ws_stats.cell(row=row_idx, column=1, value=param_name).alignment = Alignment(horizontal="left", vertical="center")
        ws_stats.cell(row=row_idx, column=2, value=round(min(values), 4)).alignment = cell_alignment
        ws_stats.cell(row=row_idx, column=3, value=round(max(values), 4)).alignment = cell_alignment
        ws_stats.cell(row=row_idx, column=4, value=round(np.mean(values), 4)).alignment = cell_alignment
        ws_stats.cell(row=row_idx, column=5, value=round(np.std(values), 4)).alignment = cell_alignment
        
        for col in range(1, 6):
            ws_stats.cell(row=row_idx, column=col).border = border
    
    # 保存Excel文件
    wb.save(output_file)
    print(f"✅ Excel表格已生成: {output_file}")
    print(f"   - 包含 5 个工作表")
    print(f"   - 20 个设备参数")
    print(f"   - 26 种调制方式")
    print(f"   - 520 个组合映射")
    

if __name__ == "__main__":
    print("="*80)
    print("生成雷达RFF数据集参数Excel表格")
    print("="*80)
    
    create_excel_with_parameters()
    
    print("\n工作表说明:")
    print("  1. 设备参数详情 - 20个设备的完整RFF参数")
    print("  2. 调制方式列表 - 26种调制类型及说明")
    print("  3. 设备调制组合映射 - 520个组合的映射关系")
    print("  4. 参数说明 - 详细的参数解释和使用说明")
    print("  5. 参数统计 - RFF参数的统计信息")
    print("="*80)

