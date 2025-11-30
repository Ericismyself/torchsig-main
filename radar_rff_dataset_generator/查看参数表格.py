"""
快速查看雷达RFF数据集参数表格内容

提供命令行方式查询Excel表格中的信息，无需打开Excel
"""

import openpyxl
import sys
from typing import Optional

def load_workbook(filename: str = "雷达RFF数据集参数表.xlsx"):
    """加载Excel工作簿"""
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        return wb
    except FileNotFoundError:
        print(f"❌ 错误: 找不到文件 '{filename}'")
        print(f"   请先运行: python 生成参数表格.py")
        sys.exit(1)

def show_device_info(device_id: Optional[int] = None):
    """显示设备参数信息"""
    wb = load_workbook()
    ws = wb["设备参数详情"]
    
    print("="*100)
    print("设备参数详情")
    print("="*100)
    
    # 打印表头
    headers = [ws.cell(1, col).value for col in range(1, 12)]
    print(f"{headers[0]:<5} {headers[1]:<20} {headers[2]:<12} {headers[3]:<12} "
          f"{headers[4]:<12} {headers[5]:<12} {headers[6]:<12}")
    print("-"*100)
    
    # 打印数据
    if device_id is not None:
        # 显示特定设备
        row = device_id + 2  # +2 因为有表头
        values = [ws.cell(row, col).value for col in range(1, 12)]
        print(f"{values[0]:<5} {values[1]:<20} {values[2]:<12.4f} {values[3]:<12.2f} "
              f"{values[4]:<12.2f} {values[5]:<12.2f} {values[6]:<12.4f}")
        
        print("\n详细参数:")
        print(f"  - IQ幅度不平衡: {values[2]:.4f}")
        print(f"  - IQ相位不平衡: {values[3]:.2f}°")
        print(f"  - 载波频偏: {values[4]:.2f} ppm")
        print(f"  - 采样率偏移: {values[5]:.2f} ppm")
        print(f"  - 相位噪声: {values[6]:.4f}")
        print(f"  - 功放平滑因子: {values[7]:.2f}")
        print(f"  - 功放饱和电平: {values[8]:.2f}")
        print(f"  - DC偏移_I: {values[9]:.4f}")
        print(f"  - DC偏移_Q: {values[10]:.4f}")
    else:
        # 显示所有设备（简化版）
        for row in range(2, 22):  # 20个设备
            values = [ws.cell(row, col).value for col in range(1, 12)]
            print(f"{values[0]:<5} {values[1]:<20} {values[2]:<12.4f} {values[3]:<12.2f} "
                  f"{values[4]:<12.2f} {values[5]:<12.2f} {values[6]:<12.4f}")
    
    print("="*100)

def show_modulation_info(mod_id: Optional[int] = None):
    """显示调制方式信息"""
    wb = load_workbook()
    ws = wb["调制方式列表"]
    
    print("="*80)
    print("调制方式列表")
    print("="*80)
    
    # 打印表头
    print(f"{'ID':<5} {'调制类型':<15} {'调制系列':<20} {'说明':<35}")
    print("-"*80)
    
    # 打印数据
    if mod_id is not None:
        # 显示特定调制
        row = mod_id + 2
        values = [ws.cell(row, col).value for col in range(1, 5)]
        print(f"{values[0]:<5} {values[1]:<15} {values[2]:<20} {values[3]:<35}")
    else:
        # 显示所有调制
        for row in range(2, 28):  # 26种调制
            values = [ws.cell(row, col).value for col in range(1, 5)]
            print(f"{values[0]:<5} {values[1]:<15} {values[2]:<20} {values[3]:<35}")
    
    print("="*80)

def show_combination_info(combined_id: int):
    """显示组合映射信息"""
    wb = load_workbook()
    ws = wb["设备调制组合映射"]
    
    print("="*80)
    print(f"组合 ID {combined_id} 详情")
    print("="*80)
    
    row = combined_id + 2  # +2 因为有表头
    
    combined_id_val = ws.cell(row, 1).value
    device_id = ws.cell(row, 2).value
    device_name = ws.cell(row, 3).value
    mod_id = ws.cell(row, 4).value
    mod_name = ws.cell(row, 5).value
    
    print(f"组合ID: {combined_id_val}")
    print(f"设备ID: {device_id} ({device_name})")
    print(f"调制ID: {mod_id} ({mod_name})")
    print("="*80)
    
    # 显示设备参数
    print("\n设备RFF参数:")
    show_device_info(device_id)
    
    # 显示调制信息
    print("\n调制类型信息:")
    show_modulation_info(mod_id)

def show_statistics():
    """显示统计信息"""
    wb = load_workbook()
    ws = wb["参数统计"]
    
    print("="*80)
    print("RFF参数统计")
    print("="*80)
    
    # 打印表头
    print(f"{'RFF参数':<25} {'最小值':<15} {'最大值':<15} {'平均值':<15} {'标准差':<15}")
    print("-"*80)
    
    # 打印数据
    for row in range(2, 11):  # 9个参数
        values = [ws.cell(row, col).value for col in range(1, 6)]
        print(f"{values[0]:<25} {values[1]:<15.4f} {values[2]:<15.4f} "
              f"{values[3]:<15.4f} {values[4]:<15.4f}")
    
    print("="*80)

def show_summary():
    """显示数据集摘要"""
    print("="*80)
    print("雷达RFF数据集摘要")
    print("="*80)
    print("数据集规模:")
    print("  - 设备数量: 20 个")
    print("  - 调制类型: 26 种")
    print("  - 组合类别: 520 个（设备-调制）")
    print("  - 每类样本数: 3,000 个")
    print("  - 总样本数: 1,560,000 个")
    print("  - 数据集大小: 约 48.6 GB")
    print()
    print("数据格式:")
    print("  - IQ样本: Complex64，形状 (2, 4096)")
    print("  - 标签: device_id (0-19) + modulation_id (0-25)")
    print("  - 存储: Sigmoid WidebandSig53 格式")
    print()
    print("Excel表格工作表:")
    print("  1. 设备参数详情 - 20个设备的完整RFF参数")
    print("  2. 调制方式列表 - 26种调制类型及说明")
    print("  3. 设备调制组合映射 - 520个组合的映射关系")
    print("  4. 参数说明 - 详细的参数解释和使用说明")
    print("  5. 参数统计 - RFF参数的统计信息")
    print("="*80)

def main():
    """主函数"""
    if len(sys.argv) == 1:
        # 无参数：显示摘要
        show_summary()
        print("\n使用方法:")
        print("  python 查看参数表格.py summary          - 显示数据集摘要")
        print("  python 查看参数表格.py devices          - 显示所有设备参数")
        print("  python 查看参数表格.py device <id>      - 显示特定设备参数")
        print("  python 查看参数表格.py modulations      - 显示所有调制类型")
        print("  python 查看参数表格.py modulation <id>  - 显示特定调制类型")
        print("  python 查看参数表格.py combination <id> - 显示组合详情")
        print("  python 查看参数表格.py stats            - 显示统计信息")
        print()
        print("示例:")
        print("  python 查看参数表格.py device 5         - 查看设备5的参数")
        print("  python 查看参数表格.py combination 100  - 查看组合ID 100")
        return
    
    command = sys.argv[1].lower()
    
    if command == "summary":
        show_summary()
    elif command == "devices":
        show_device_info()
    elif command == "device":
        if len(sys.argv) < 3:
            print("❌ 错误: 请指定设备ID")
            print("   用法: python 查看参数表格.py device <id>")
            sys.exit(1)
        device_id = int(sys.argv[2])
        if not 0 <= device_id < 20:
            print(f"❌ 错误: 设备ID必须在 0-19 之间")
            sys.exit(1)
        show_device_info(device_id)
    elif command == "modulations":
        show_modulation_info()
    elif command == "modulation":
        if len(sys.argv) < 3:
            print("❌ 错误: 请指定调制ID")
            print("   用法: python 查看参数表格.py modulation <id>")
            sys.exit(1)
        mod_id = int(sys.argv[2])
        if not 0 <= mod_id < 26:
            print(f"❌ 错误: 调制ID必须在 0-25 之间")
            sys.exit(1)
        show_modulation_info(mod_id)
    elif command == "combination":
        if len(sys.argv) < 3:
            print("❌ 错误: 请指定组合ID")
            print("   用法: python 查看参数表格.py combination <id>")
            sys.exit(1)
        combined_id = int(sys.argv[2])
        if not 0 <= combined_id < 520:
            print(f"❌ 错误: 组合ID必须在 0-519 之间")
            sys.exit(1)
        show_combination_info(combined_id)
    elif command == "stats":
        show_statistics()
    else:
        print(f"❌ 错误: 未知命令 '{command}'")
        print("\n可用命令:")
        print("  summary, devices, device, modulations, modulation, combination, stats")
        sys.exit(1)

if __name__ == "__main__":
    main()

